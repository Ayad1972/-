# -*- coding: utf-8 -*-
"""
نقل أرقام الموظفين والخدمة من Excel إلى جدول FoxPro DBF
مع الحفاظ على نفس الترميز والرموز (متوافق مع أنظمة الفوكس القديمة).

الاستخدام:
  python update_mfile_from_excel.py
  python update_mfile_from_excel.py --excel "data\\employees.xlsx" --dbf "MFILE.DBF"
  python update_mfile_from_excel.py --inspect-only
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import shutil
import struct
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    import openpyxl
except ImportError:
    print("المكتبة openpyxl غير مثبتة. نفّذ: pip install openpyxl")
    sys.exit(1)

try:
    from dbfread import DBF
except ImportError:
    print("المكتبة dbfread غير مثبتة. نفّذ: pip install dbfread")
    sys.exit(1)

try:
    import dbf as dbf_lib
except ImportError:
    dbf_lib = None


# معرفات لغة FoxPro / dBASE الشائعة
CODEPAGE_MAP = {
    0x01: "cp437",
    0x02: "cp850",
    0x03: "cp1252",
    0x57: "cp1252",
    0x64: "cp852",
    0x65: "cp866",
    0x66: "cp865",
    0x67: "cp861",
    0x6A: "cp737",
    0x6B: "cp857",
    0x7D: "cp1255",
    0x7E: "cp1256",  # Arabic Windows - الأهم للفوكس العربي
    0xC8: "cp1250",
    0xC9: "cp1251",
    0xCA: "cp1254",
    0xCB: "cp1253",
}

# أسماء أعمدة محتملة في Excel / DBF
EMP_ALIASES = [
    "رقم الموظف",
    "رقمالموظف",
    "رقم_الموظف",
    "الموظف",
    "EMPNO",
    "EMP_NO",
    "PNO",
    "EMPLOYEE",
    "EMP",
    "NO",
    "NUM",
    "NUMBER",
    "ID",
    "كود",
    "الكود",
    "CODE",
]

SERVICE_ALIASES = [
    "الخدمة",
    "خدمة",
    "SERVICE",
    "SERV",
    "JOB",
    "الوظيفة",
    "وظيفة",
    "WORK",
    "DEPT",
    "القسم",
    "قسم",
    "KHIDMA",
    "KHEDMA",
]


def normalize_name(name: Any) -> str:
    if name is None:
        return ""
    text = str(name).strip().upper()
    text = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    text = text.replace("ة", "ه").replace("_", "").replace(" ", "").replace("-", "")
    return text


def find_column(headers: Sequence[Any], aliases: Sequence[str]) -> Optional[int]:
    norm_headers = [normalize_name(h) for h in headers]
    norm_aliases = [normalize_name(a) for a in aliases]
    for i, h in enumerate(norm_headers):
        if not h:
            continue
        for a in norm_aliases:
            if h == a or a in h or h in a:
                return i
    return None


def find_dbf_field(field_names: Sequence[str], aliases: Sequence[str]) -> Optional[str]:
    norm_fields = {normalize_name(f): f for f in field_names}
    for alias in aliases:
        na = normalize_name(alias)
        if na in norm_fields:
            return norm_fields[na]
        for nf, original in norm_fields.items():
            if na and (na in nf or nf in na):
                return original
    return None


def read_dbf_language_driver(dbf_path: Path) -> int:
    with dbf_path.open("rb") as f:
        header = f.read(32)
    if len(header) < 30:
        raise ValueError("ملف DBF تالف أو غير مكتمل")
    return header[29]


def detect_encoding(dbf_path: Path, forced: Optional[str] = None) -> Tuple[str, int]:
    ldid = read_dbf_language_driver(dbf_path)
    if forced:
        return forced, ldid
    enc = CODEPAGE_MAP.get(ldid)
    if enc:
        return enc, ldid
    # أنظمة فوكس عربية قديمة غالباً تستخدم 1256 حتى لو LDID=0
    return "cp1256", ldid


def sanitize_fox_text(value: Any, encoding: str, width: Optional[int] = None) -> str:
    """تحويل النص إلى رموز متوافقة مع ترميز جدول DBF القديم."""
    if value is None:
        text = ""
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()

    # توحيد المسافات وإزالة محارف غير قابلة للتمثيل في الترميز الهدف
    text = re.sub(r"\s+", " ", text)
    try:
        text.encode(encoding)
    except UnicodeEncodeError:
        text = text.encode(encoding, errors="replace").decode(encoding)

    if width is not None and width > 0:
        # حقول Character في DBF بعرض ثابت
        encoded = text.encode(encoding, errors="replace")
        if len(encoded) > width:
            encoded = encoded[:width]
            text = encoded.decode(encoding, errors="ignore")
    return text


def sanitize_emp_no(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def load_excel_rows(
    excel_path: Path,
    emp_col: Optional[str] = None,
    service_col: Optional[str] = None,
) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("ملف Excel فارغ")

    headers = list(rows[0])
    # إذا الصف الأول ليس عناوين، نفترض ترتيب: رقم الموظف، الخدمة
    looks_like_header = any(isinstance(h, str) and h.strip() for h in headers)

    meta: Dict[str, Any] = {"sheet": ws.title, "headers": headers}

    if looks_like_header:
        if emp_col:
            emp_idx = next((i for i, h in enumerate(headers) if str(h).strip() == emp_col), None)
        else:
            emp_idx = find_column(headers, EMP_ALIASES)

        if service_col:
            svc_idx = next((i for i, h in enumerate(headers) if str(h).strip() == service_col), None)
        else:
            svc_idx = find_column(headers, SERVICE_ALIASES)

        data_rows = rows[1:]
    else:
        emp_idx, svc_idx = 0, 1 if len(headers) > 1 else None
        data_rows = rows
        meta["headers"] = ["COL1", "COL2"]

    if emp_idx is None:
        raise ValueError(
            "تعذر إيجاد عمود رقم الموظف في Excel. "
            "استخدم --emp-excel-col أو تأكد من اسم العمود."
        )
    if svc_idx is None:
        raise ValueError(
            "تعذر إيجاد عمود الخدمة في Excel. "
            "استخدم --service-excel-col أو تأكد من اسم العمود."
        )

    meta["emp_excel_col"] = headers[emp_idx] if looks_like_header else f"COL{emp_idx+1}"
    meta["service_excel_col"] = headers[svc_idx] if looks_like_header else f"COL{svc_idx+1}"

    out: List[Dict[str, str]] = []
    for row in data_rows:
        if row is None:
            continue
        emp = sanitize_emp_no(row[emp_idx] if emp_idx < len(row) else None)
        svc_raw = row[svc_idx] if svc_idx < len(row) else None
        if not emp and (svc_raw is None or str(svc_raw).strip() == ""):
            continue
        out.append({"emp_no": emp, "service": "" if svc_raw is None else str(svc_raw).strip()})
    return out, meta


def inspect_files(excel_path: Path, dbf_path: Path) -> int:
    print("=" * 60)
    print("فحص الملفات")
    print("=" * 60)

    print(f"\nExcel: {excel_path}")
    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        print(f"  الورقة: {ws.title}")
        print(f"  عدد الصفوف: {len(rows)}")
        if rows:
            print(f"  الصف الأول: {list(rows[0])}")
            if len(rows) > 1:
                print(f"  صف نموذجي: {list(rows[1])}")
    else:
        print("  الملف غير موجود")

    print(f"\nDBF: {dbf_path}")
    if dbf_path.exists():
        ldid = read_dbf_language_driver(dbf_path)
        enc, _ = detect_encoding(dbf_path)
        print(f"  Language Driver ID: {ldid} (0x{ldid:02X})")
        print(f"  الترميز المقترح: {enc}")
        table = DBF(str(dbf_path), encoding=enc, ignore_missing_memofile=True, load=True)
        records = list(getattr(table, "records", table))
        print(f"  الحقول: {table.field_names}")
        print(f"  عدد السجلات: {len(records)}")
        if records:
            print(f"  سجل نموذجي: {dict(records[0])}")
        emp_field = find_dbf_field(table.field_names, EMP_ALIASES)
        svc_field = find_dbf_field(table.field_names, SERVICE_ALIASES)
        print(f"  حقل رقم الموظف المكتشف: {emp_field}")
        print(f"  حقل الخدمة المكتشف: {svc_field}")
    else:
        print("  الملف غير موجود")
        print("  ضع MFILE.DBF بجانب البرنامج أو في مجلد data")
    if not excel_path.exists():
        print("  ضع ملف Excel في مجلد المشروع أو data")
    return 0 if dbf_path.exists() else 1


def field_info(table: "dbf_lib.Table", field_name: str) -> Tuple[str, int, int]:
    """يرجع (type, length, decimal)."""
    spec = table.field_info(field_name)
    # dbf library returns namedtuple-like values depending on version
    if hasattr(spec, "field_type"):
        return str(spec.field_type), int(spec.length), int(getattr(spec, "decimal_count", 0) or 0)
    if isinstance(spec, (tuple, list)) and len(spec) >= 2:
        return str(spec[0]), int(spec[1]), int(spec[2] if len(spec) > 2 else 0)
    # fallback
    return "C", 20, 0


def coerce_for_field(value: str, ftype: str, length: int, decimals: int, encoding: str) -> Any:
    ftype = ftype.upper()
    if ftype in ("C", "V", "W"):
        return sanitize_fox_text(value, encoding, width=length)
    if ftype == "N":
        if value.strip() == "":
            return 0 if decimals == 0 else 0.0
        try:
            return int(float(value)) if decimals == 0 else float(value)
        except ValueError:
            return 0 if decimals == 0 else 0.0
    if ftype == "F":
        try:
            return float(value)
        except ValueError:
            return 0.0
    if ftype == "L":
        return str(value).strip().lower() in ("1", "y", "t", "true", "yes", "نعم")
    if ftype == "D":
        # نترك التاريخ كما هو إن لم يكن تاريخاً واضحاً
        return None
    return sanitize_fox_text(value, encoding, width=length if length else None)


def update_dbf(
    excel_path: Path,
    dbf_path: Path,
    encoding_forced: Optional[str] = None,
    emp_excel_col: Optional[str] = None,
    service_excel_col: Optional[str] = None,
    emp_dbf_field: Optional[str] = None,
    service_dbf_field: Optional[str] = None,
    mode: str = "upsert",
    dry_run: bool = False,
) -> None:
    if dbf_lib is None:
        raise RuntimeError("المكتبة dbf غير مثبتة. نفّذ: pip install dbf")

    if not excel_path.exists():
        raise FileNotFoundError(f"ملف Excel غير موجود: {excel_path}")
    if not dbf_path.exists():
        raise FileNotFoundError(f"ملف DBF غير موجود: {dbf_path}")

    encoding, ldid = detect_encoding(dbf_path, encoding_forced)
    excel_rows, excel_meta = load_excel_rows(excel_path, emp_excel_col, service_excel_col)

    # قراءة الحقول عبر dbfread أولاً للاكتشاف
    preview = DBF(str(dbf_path), encoding=encoding, ignore_missing_memofile=True, load=True)
    fields = list(preview.field_names)

    emp_field = emp_dbf_field or find_dbf_field(fields, EMP_ALIASES)
    svc_field = service_dbf_field or find_dbf_field(fields, SERVICE_ALIASES)

    if not emp_field:
        raise ValueError(
            f"تعذر إيجاد حقل رقم الموظف في DBF. الحقول المتاحة: {fields}. "
            "مرّر الاسم عبر --emp-dbf-field"
        )
    if not svc_field:
        raise ValueError(
            f"تعذر إيجاد حقل الخدمة في DBF. الحقول المتاحة: {fields}. "
            "مرّر الاسم عبر --service-dbf-field"
        )

    print("=" * 60)
    print("خطة النقل")
    print("=" * 60)
    print(f"Excel عمود الموظف : {excel_meta['emp_excel_col']}")
    print(f"Excel عمود الخدمة : {excel_meta['service_excel_col']}")
    print(f"DBF حقل الموظف    : {emp_field}")
    print(f"DBF حقل الخدمة    : {svc_field}")
    print(f"Language Driver   : {ldid} (0x{ldid:02X})")
    print(f"الترميز المستخدم  : {encoding}")
    print(f"عدد صفوف Excel    : {len(excel_rows)}")
    print(f"وضع العمل         : {mode}")
    print(f"تجربة فقط         : {dry_run}")

    backup_path = dbf_path.with_suffix(
        dbf_path.suffix + f".bak_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )
    if not dry_run:
        shutil.copy2(dbf_path, backup_path)
        # نسخة من ملف المذكرة إن وجد
        for memo_ext in (".FPT", ".fpt", ".DBT", ".dbt"):
            memo = dbf_path.with_suffix(memo_ext)
            if memo.exists():
                shutil.copy2(memo, Path(str(backup_path) + memo_ext))
        print(f"تم إنشاء نسخة احتياطية: {backup_path}")

    # فتح الجدول للكتابة بنفس الترميز
    table = dbf_lib.Table(str(dbf_path), codepage=encoding, ignore_memos=True)
    table.open(mode=dbf_lib.READ_WRITE)

    try:
        emp_type, emp_len, emp_dec = field_info(table, emp_field)
        svc_type, svc_len, svc_dec = field_info(table, svc_field)

        # فهرسة السجلات الحالية حسب رقم الموظف
        index: Dict[str, List[Any]] = {}
        for rec in table:
            key = sanitize_emp_no(rec[emp_field])
            if key:
                index.setdefault(key, []).append(rec)

        updated = 0
        inserted = 0
        skipped = 0

        for item in excel_rows:
            emp_no = sanitize_emp_no(item["emp_no"])
            service = sanitize_fox_text(item["service"], encoding, width=svc_len if svc_type.upper() in ("C", "V", "W") else None)

            if not emp_no:
                skipped += 1
                continue

            emp_value = coerce_for_field(emp_no, emp_type, emp_len, emp_dec, encoding)
            svc_value = coerce_for_field(service, svc_type, svc_len, svc_dec, encoding)

            if mode in ("upsert", "update") and emp_no in index:
                for rec in index[emp_no]:
                    if dry_run:
                        print(f"[UPDATE] {emp_field}={emp_value!r} {svc_field}={svc_value!r}")
                    else:
                        dbf_lib.write(
                            rec,
                            **{
                                emp_field: emp_value,
                                svc_field: svc_value,
                            },
                        )
                    updated += 1
            elif mode in ("upsert", "append"):
                if dry_run:
                    print(f"[INSERT] {emp_field}={emp_value!r} {svc_field}={svc_value!r}")
                else:
                    data = {name: None for name in table.field_names}
                    data[emp_field] = emp_value
                    data[svc_field] = svc_value
                    # تعبئة الحقول النصية الفارغة بمسافات آمنة بدل None إن لزم
                    for name in table.field_names:
                        ftype, flen, fdec = field_info(table, name)
                        if data[name] is None and ftype.upper() in ("C", "V", "W"):
                            data[name] = ""
                        if data[name] is None and ftype.upper() in ("N", "F"):
                            data[name] = 0 if fdec == 0 and ftype.upper() == "N" else 0.0
                        if data[name] is None and ftype.upper() == "L":
                            data[name] = False
                    table.append(data)
                inserted += 1
                # تحديث الفهرس حتى لا يتكرر الإدراج لنفس الرقم ضمن نفس التشغيل
                index.setdefault(emp_no, []).append(None)
            else:
                skipped += 1

        # الحفاظ على Language Driver الأصلي في الهيدر
        if not dry_run:
            table.close()
            preserve_language_driver(dbf_path, ldid)
            # إعادة الفتح ليست مطلوبة بعد الحفظ
        else:
            table.close()

        print("-" * 60)
        print(f"تم التحديث : {updated}")
        print(f"تم الإدراج : {inserted}")
        print(f"تم التجاوز : {skipped}")
        if not dry_run:
            print("اكتمل النقل مع الحفاظ على ترميز/رموز جدول DBF.")
            print("افتح MFILE.DBF في نظام الفوكس وتحقق من الحقول.")
    finally:
        if table.status != dbf_lib.CLOSED:
            table.close()


def preserve_language_driver(dbf_path: Path, ldid: int) -> None:
    """إعادة كتابة بايت لغة الهيدر كما كان، لضمان توافق الفوكس القديم."""
    with dbf_path.open("r+b") as f:
        f.seek(29)
        current = f.read(1)
        if not current:
            return
        if current[0] != ldid:
            f.seek(29)
            f.write(struct.pack("B", ldid))


def default_paths() -> Tuple[Path, Path]:
    try:
        from afrad_portable import default_update_paths

        return default_update_paths()
    except Exception:
        root = Path(__file__).resolve().parent
        candidates_excel = [
            root / "data" / "New Microsoft Excel Worksheet.xlsx",
            root / "New Microsoft Excel Worksheet.xlsx",
        ]
        candidates_dbf = [
            root / "MFILE.DBF",
            root / "data" / "MFILE.DBF",
            root / "MFILE_updated.DBF",
        ]
        excel = next((p for p in candidates_excel if p.exists()), candidates_excel[0])
        dbfp = next((p for p in candidates_dbf if p.exists()), candidates_dbf[0])
        return excel, dbfp


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    excel_default, dbf_default = default_paths()
    p = argparse.ArgumentParser(description="نقل أرقام الموظفين والخدمة من Excel إلى MFILE.DBF")
    p.add_argument("--excel", default=str(excel_default), help="مسار ملف Excel")
    p.add_argument("--dbf", default=str(dbf_default), help="مسار ملف DBF")
    p.add_argument("--encoding", default=None, help="فرض ترميز مثل cp1256")
    p.add_argument("--emp-excel-col", default=None, help="اسم عمود رقم الموظف في Excel")
    p.add_argument("--service-excel-col", default=None, help="اسم عمود الخدمة في Excel")
    p.add_argument("--emp-dbf-field", default=None, help="اسم حقل رقم الموظف في DBF")
    p.add_argument("--service-dbf-field", default=None, help="اسم حقل الخدمة في DBF")
    p.add_argument(
        "--mode",
        choices=["upsert", "update", "append"],
        default="upsert",
        help="upsert=تحديث إن وجد وإلا إدراج | update=تحديث فقط | append=إدراج دائماً",
    )
    p.add_argument("--inspect-only", action="store_true", help="عرض البنية فقط دون تعديل")
    p.add_argument("--dry-run", action="store_true", help="تجربة بدون كتابة على DBF")
    return p.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    excel_path = Path(args.excel)
    dbf_path = Path(args.dbf)

    # على Windows قد تكون المسارات بحروف كبيرة/صغيرة مختلفة
    if os.name == "nt":
        excel_path = Path(os.path.expandvars(str(excel_path)))
        dbf_path = Path(os.path.expandvars(str(dbf_path)))

    if args.inspect_only:
        return inspect_files(excel_path, dbf_path)

    try:
        update_dbf(
            excel_path=excel_path,
            dbf_path=dbf_path,
            encoding_forced=args.encoding,
            emp_excel_col=args.emp_excel_col,
            service_excel_col=args.service_excel_col,
            emp_dbf_field=args.emp_dbf_field,
            service_dbf_field=args.service_dbf_field,
            mode=args.mode,
            dry_run=args.dry_run,
        )
    except FileNotFoundError as exc:
        print("\nالملف غير موجود على هذه الحاسبة.")
        print(exc)
        print("انسخ الملفات إلى مجلد المشروع أو مجلد data ثم شغّل START.bat")
        return 1
    except Exception as exc:
        print("\nتعذر النقل:", exc)
        print("ضع Excel و MFILE.DBF في مجلد البرنامج أو في data ثم أعد التشغيل.")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
