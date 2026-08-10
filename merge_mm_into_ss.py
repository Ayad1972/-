# -*- coding: utf-8 -*-
"""
دمج أرقام mm.xlsx في ss.xls حسب الاسم المشترك.
يحفظ الناتج في H:\\ss_updated.xls (أو .xlsx إن تعذر xls).
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl


NAME_ALIASES = [
    "الاسم", "اسم", "اسم الموظف", "اسم_الموظف", "الموظف",
    "NAME", "EMP_NAME", "EMPLOYEE", "FULLNAME", "FULL_NAME",
]
NUMBER_ALIASES = [
    "الرقم", "رقم", "رقم الموظف", "رقم_الموظف", "رقم وظيفي",
    "NUMBER", "NO", "NUM", "EMPNO", "EMP_NO", "ID", "CODE", "الكود", "كود",
]


def normalize_name(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    text = text.replace("ة", "ه").replace("ى", "ي")
    text = re.sub(r"\s+", " ", text)
    return text.lower()


def normalize_header(value: Any) -> str:
    return normalize_name(value).replace(" ", "").replace("_", "").replace("-", "")


def find_col(headers: Sequence[Any], aliases: Sequence[str]) -> Optional[int]:
    norms = [normalize_header(h) for h in headers]
    alias_norms = [normalize_header(a) for a in aliases]
    for i, h in enumerate(norms):
        if not h:
            continue
        for a in alias_norms:
            if h == a or a in h or h in a:
                return i
    return None


def read_xls(path: Path) -> Tuple[List[Any], List[List[Any]]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("لقراءة .xls ثبّت: pip install xlrd==1.2.0") from exc

    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    rows = [sheet.row_values(r) for r in range(sheet.nrows)]
    if not rows:
        return [], []
    return list(rows[0]), [list(r) for r in rows[1:]]


def read_xlsx(path: Path) -> Tuple[List[Any], List[List[Any]]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return list(rows[0]), [list(r) for r in rows[1:]]


def read_any(path: Path) -> Tuple[List[Any], List[List[Any]]]:
    if path.suffix.lower() == ".xls":
        return read_xls(path)
    return read_xlsx(path)


def detect_columns(headers: Sequence[Any], sample_rows: Sequence[Sequence[Any]]) -> Tuple[int, int, bool]:
    name_idx = find_col(headers, NAME_ALIASES)
    num_idx = find_col(headers, NUMBER_ALIASES)
    if name_idx is not None and num_idx is not None:
        return name_idx, num_idx, True

    # بدون عناوين واضحة: عمود 1 اسم، عمود 2 رقم
    return 0, 1, False


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def main(argv: Optional[Sequence[str]] = None) -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--ss", default=r"H:\ss.xls")
    p.add_argument("--mm", default=r"H:\mm.xlsx")
    p.add_argument("--out", default=r"H:\ss_updated.xlsx")
    args = p.parse_args(argv)

    ss_path = Path(args.ss)
    mm_path = Path(args.mm)
    out_path = Path(args.out)

    if not ss_path.exists():
        print("الملف غير موجود:", ss_path)
        return 1
    if not mm_path.exists():
        print("الملف غير موجود:", mm_path)
        return 1

    ss_headers, ss_rows = read_any(ss_path)
    mm_headers, mm_rows = read_any(mm_path)

    # إذا ss بلا صف عناوين واضح، أعد صف العناوين كبيانات
    ss_name_idx, ss_num_idx, ss_has_header = detect_columns(ss_headers, ss_rows)
    mm_name_idx, mm_num_idx, mm_has_header = detect_columns(mm_headers, mm_rows)

    if not ss_has_header:
        ss_rows = [list(ss_headers)] + ss_rows
        ss_headers = ["الاسم", "الرقم"] + [f"COL{i}" for i in range(3, max(len(ss_headers), 2) + 1)]
        ss_name_idx, ss_num_idx = 0, 1
    if not mm_has_header:
        mm_rows = [list(mm_headers)] + mm_rows
        mm_headers = ["الاسم", "الرقم"]
        mm_name_idx, mm_num_idx = 0, 1

    print("ss:", ss_headers[ss_name_idx], "/", ss_headers[ss_num_idx])
    print("mm:", mm_headers[mm_name_idx], "/", mm_headers[mm_num_idx])

    mm_map: Dict[str, str] = {}
    for row in mm_rows:
        if mm_name_idx >= len(row) or mm_num_idx >= len(row):
            continue
        key = normalize_name(row[mm_name_idx])
        num = cell_to_text(row[mm_num_idx])
        if key and num and key not in mm_map:
            mm_map[key] = num

    updated = matched_same = not_found = 0
    for row in ss_rows:
        while len(row) <= max(ss_name_idx, ss_num_idx):
            row.append(None)
        key = normalize_name(row[ss_name_idx])
        if not key:
            continue
        if key in mm_map:
            new_num = mm_map[key]
            old_num = cell_to_text(row[ss_num_idx])
            if old_num != new_num:
                row[ss_num_idx] = new_num
                updated += 1
                print(f"تحديث: {row[ss_name_idx]} | من [{old_num}] إلى [{new_num}]")
            else:
                matched_same += 1
        else:
            not_found += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(ss_headers)
    for row in ss_rows:
        ws.append(row)
    wb.save(out_path)

    print("--------------------------------------------")
    print("تم التحديث          :", updated)
    print("متطابق بدون تغيير   :", matched_same)
    print("اسم غير موجود في mm :", not_found)
    print("الملف الناتج        :", out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
